"""
services/lease_tradeout_parser.py
Parses a Yardi-style "Lease Trade-Out" export (.xlsx).

Expected structure (Yardi v2.2):
  Row 2: Title "Lease Trade-out"
  Row 3: Property name
  Row 4: Date range
  Row 6: Section header "Summary"
  Row 7: Summary table header (Current Lease Type, Leases, SQFT, Days Vacant,
         Prior Lease Rent, Prior Lease Rent/SQFT, Prior Effective Rent,
         Prior Effective Rent/SQFT, Current Lease Rent, Current Lease Rent/SQFT,
         Current Effective Rent, Current Effective Rent/SQFT,
         Lease Rent Change ($), Lease Rent Change (%),
         Effective Rent Change ($), Effective Rent Change (%))
  Row 8+: Summary rows by Current Lease Type
          ("Application", "Renewal", "Month To Month", "Total/Average:")
  Row 13: "Detail" section header
  Row 14: Detail header (Skip MTM?, Unit Type, Bldg-Unit, SQFT, Days Vacant,
          Considered Rented On, Prior Resident, Current Resident, ...)
  Row 15+: Per-unit detail rows.

Returns a dict with parsed summary, detail rows, and rolled-up KPIs.
"""
from datetime import datetime
from openpyxl import load_workbook


def _safe_float(v):
    if v is None or v == "" or isinstance(v, bool):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _safe_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_lease_tradeout(file_path: str) -> dict:
    """
    Parse a Lease Trade-Out file.

    Returns
    -------
    {
        "property": str | None,
        "period":   str | None,          # "04/01/2026 - 04/30/2026"
        "summary": [                     # list of summary rows by lease type
            {
                "lease_type": str,
                "leases": int,
                "avg_sqft": float|None,
                "avg_days_vacant": float|None,
                "prior_lease_rent": float|None,
                "current_lease_rent": float|None,
                "lease_rent_change_dollar": float|None,
                "lease_rent_change_pct": float|None,
                "effective_rent_change_dollar": float|None,
                "effective_rent_change_pct": float|None,
            }, ...
        ],
        "total": { ... same shape as summary entry, lease_type="Total" ... },
        "detail": [                      # per-unit lease activity
            {
                "unit_type": str,
                "bldg_unit": str,
                "sqft": float|None,
                "days_vacant": float|None,
                "considered_rented_on": datetime|None,
                "prior_resident": str|None,
                "current_resident": str|None,
                "current_lease_type": str|None,   # Application / Renewal / MTM
                "prior_lease_rent": float|None,
                "current_lease_rent": float|None,
                "lease_rent_change_dollar": float|None,
                "lease_rent_change_pct": float|None,
            }, ...
        ],
        "errors":   [str],
        "warnings": [str],
    }
    """
    errors, warnings = [], []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return {"errors": [f"Cannot open file: {e}"], "warnings": []}

    # Property data is in the first sheet (named after the property);
    # Report Parameters is the second sheet — skip it.
    sheet_names = [n for n in wb.sheetnames if n != "Report Parameters"]
    if not sheet_names:
        return {"errors": ["No data sheet found."], "warnings": []}
    ws = wb[sheet_names[0]]
    rows = list(ws.iter_rows(values_only=True))

    # ── Header metadata ────────────────────────────────────────────────────
    property_name = None
    period_str = None

    for i, r in enumerate(rows[:6]):
        if r and r[0] and isinstance(r[0], str):
            s = r[0].strip()
            if "Lease Trade" in s:
                continue
            if property_name is None and s and i == 2:
                property_name = s
            elif period_str is None and " - " in s and "/" in s and i == 3:
                period_str = s

    # ── Locate Summary header row ──────────────────────────────────────────
    summary_header_idx = None
    for i, r in enumerate(rows):
        if not r:
            continue
        first = r[0]
        if isinstance(first, str) and first.strip() == "Current Lease Type":
            summary_header_idx = i
            break

    summary_rows = []
    total_row = None
    if summary_header_idx is None:
        warnings.append("Could not locate Summary section.")
    else:
        # Column indices in summary table
        # 0 Current Lease Type | 1 Leases | 2 SQFT | 3 Days Vacant
        # 4 Prior Lease Rent   | 5 Prior Lease Rent/SQFT
        # 6 Prior Effective Rent | 7 Prior Effective Rent/SQFT
        # 8 Current Lease Rent | 9 Current Lease Rent/SQFT
        # 10 Current Effective Rent | 11 Current Effective Rent/SQFT
        # 12 Lease Rent Change $ | 13 Lease Rent Change %
        # 14 Effective Rent Change $ | 15 Effective Rent Change %
        for r in rows[summary_header_idx + 1:]:
            if not r or not r[0]:
                # Blank row ends the summary block
                break
            lease_type = _safe_str(r[0])
            if not lease_type:
                break
            entry = {
                "lease_type":                      lease_type.rstrip(":").strip(),
                "leases":                          int(_safe_float(r[1]) or 0),
                "avg_sqft":                        _safe_float(r[2]),
                "avg_days_vacant":                 _safe_float(r[3]),
                "prior_lease_rent":                _safe_float(r[4]),
                "current_lease_rent":              _safe_float(r[8]) if len(r) > 8 else None,
                "lease_rent_change_dollar":        _safe_float(r[12]) if len(r) > 12 else None,
                "lease_rent_change_pct":           _safe_float(r[13]) if len(r) > 13 else None,
                "effective_rent_change_dollar":    _safe_float(r[14]) if len(r) > 14 else None,
                "effective_rent_change_pct":       _safe_float(r[15]) if len(r) > 15 else None,
            }
            if "total" in lease_type.lower():
                total_row = entry
            else:
                summary_rows.append(entry)

    # ── Locate Detail header row ───────────────────────────────────────────
    detail_header_idx = None
    for i, r in enumerate(rows):
        if not r:
            continue
        # Detail header has "Bldg-Unit" in column index 2
        if len(r) > 2 and isinstance(r[2], str) and r[2].strip() == "Bldg-Unit":
            detail_header_idx = i
            break

    detail = []
    if detail_header_idx is None:
        warnings.append("Could not locate Detail section.")
    else:
        # Detail column layout (0-indexed):
        # 0 Skip MTM? | 1 Unit Type | 2 Bldg-Unit | 3 SQFT | 4 Days Vacant
        # 5 Considered Rented On | 6 Prior Resident | 7 Current Resident
        # 8 Current Lease Type | 9 Prior Lease Term | 10 Current Lease Term
        # 11 Prior Lease Rent | 12 Prior Lease Rent/SQFT
        # 13 Prior Effective Rent | 14 Prior Effective Rent/SQFT
        # 15 Current Lease Rent | 16 Current Lease Rent/SQFT
        # 17 Current Effective Rent | 18 Current Effective Rent/SQFT
        # (further columns: rent change $, % — depends on report version)
        for r in rows[detail_header_idx + 1:]:
            if not r or not _safe_str(r[2] if len(r) > 2 else None):
                continue
            prior_rent   = _safe_float(r[11]) if len(r) > 11 else None
            current_rent = _safe_float(r[15]) if len(r) > 15 else None
            rent_chg_d = (current_rent - prior_rent
                          if (prior_rent is not None and current_rent is not None)
                          else None)
            rent_chg_p = (rent_chg_d / prior_rent
                          if (rent_chg_d is not None and prior_rent not in (None, 0))
                          else None)
            detail.append({
                "unit_type":              _safe_str(r[1]) if len(r) > 1 else None,
                "bldg_unit":              _safe_str(r[2]) if len(r) > 2 else None,
                "sqft":                   _safe_float(r[3]) if len(r) > 3 else None,
                "days_vacant":            _safe_float(r[4]) if len(r) > 4 else None,
                "considered_rented_on":   r[5] if (len(r) > 5 and isinstance(r[5], datetime)) else None,
                "prior_resident":         _safe_str(r[6]) if len(r) > 6 else None,
                "current_resident":       _safe_str(r[7]) if len(r) > 7 else None,
                "current_lease_type":     _safe_str(r[8]) if len(r) > 8 else None,
                "prior_lease_rent":       prior_rent,
                "current_lease_rent":     current_rent,
                "lease_rent_change_dollar": rent_chg_d,
                "lease_rent_change_pct":  rent_chg_p,
            })

    if not summary_rows and not detail:
        errors.append("Lease Trade-Out file contains neither summary nor detail rows.")

    return {
        "property":  property_name,
        "period":    period_str,
        "summary":   summary_rows,
        "total":     total_row,
        "detail":    detail,
        "errors":    errors,
        "warnings":  warnings,
    }
