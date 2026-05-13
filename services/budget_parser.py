"""
services/budget_parser.py
Parses a standardized Budget Template (.xlsx).

The Budget template is structurally identical to the T12 template:
  Row 4  → "Budget As Of Date:" in col C, date in col D
  Row 8  → header row: col C = "Category", col D = "Budget Line-Item Name",
           cols E-P = monthly dates (12 months),
           col R = F12 (full-year forecast), col S = F6, col T = F3, col U = F1

The output dict mirrors `parse_t12()` so the financials view can consume
budget_data and t12_data with the same row-matching logic.
"""
from datetime import datetime
from openpyxl import load_workbook

# Re-use the same classification keywords as the T12 parser so categorisation
# is consistent across both files.
from services.t12_parser import (
    REVENUE_KEYWORDS, EXPENSE_KEYWORDS, NOI_KEYWORDS,
    _is_revenue, _is_expense, _is_noi, _safe_float,
)


def parse_budget(file_path: str) -> dict:
    """
    Parse a Budget Template (.xlsx) and return a structured dict.

    Returns the same shape as `parse_t12()`. The T12/T6/T3/T1 keys are populated
    with the budget's F12/F6/F3/F1 values so downstream code (financials view,
    row matching) doesn't need to special-case anything.

    Additional keys for budget-specific consumers:
      - "monthly_revenue": [float, ...12]   (alias of monthly_totals.revenue)
      - "annual_revenue":  float            (= total_revenue_t12)
    """
    errors, warnings = [], []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return {"errors": [f"Cannot open file: {e}"], "warnings": []}

    # Template uses sheet name "Budget"; fall back to first sheet.
    if "Budget" in wb.sheetnames:
        ws = wb["Budget"]
    else:
        ws = wb.active
        warnings.append(f"Sheet 'Budget' not found — using '{ws.title}'.")

    all_rows = list(ws.iter_rows(values_only=True))

    # ── Step 1: Find as-of date and header row ────────────────────────────
    as_of_date = None
    header_row_idx = None
    cat_col = item_col = None
    month_col_indices = []
    month_dates = []

    for i, row in enumerate(all_rows):
        for j, cell in enumerate(row):
            # As-of date
            if isinstance(cell, str) and "Budget As Of Date" in cell:
                for k in range(j + 1, min(j + 5, len(row))):
                    if isinstance(row[k], datetime):
                        as_of_date = row[k]
                        break
            # Header row marker — "Category" cell with line-item header to the right
            if isinstance(cell, str) and cell.strip() == "Category":
                cat_col = j
                for k in range(j + 1, min(j + 5, len(row))):
                    if isinstance(row[k], str) and "line" in row[k].lower():
                        item_col = k
                        break
                # Monthly date columns after item_col
                month_cols = []
                for k in range((item_col or j) + 1, len(row)):
                    if isinstance(row[k], datetime):
                        month_cols.append((k, row[k]))
                if month_cols:
                    header_row_idx = i
                    month_dates = [d for _, d in month_cols[:12]]
                    month_col_indices = [c for c, _ in month_cols[:12]]
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        errors.append(
            "Could not locate header row in budget file "
            "(expected 'Category' column with monthly date columns after it)."
        )
        return {"errors": errors, "warnings": warnings}

    if as_of_date is None:
        warnings.append("Budget as-of date not found; using last monthly column date.")
        as_of_date = month_dates[-1] if month_dates else None

    # ── Step 2: Locate F12/F6/F3/F1 columns ────────────────────────────────
    f12_col = f6_col = f3_col = f1_col = conf_col = None
    hdr = all_rows[header_row_idx]
    for j, cell in enumerate(hdr):
        if not isinstance(cell, str):
            continue
        cl = cell.strip().upper()
        if cl in ("F12", "T12") and f12_col is None and j > month_col_indices[-1]:
            f12_col = j
        elif cl in ("F6", "T6") and f6_col is None and j > month_col_indices[-1]:
            f6_col = j
        elif cl in ("F3", "T3") and f3_col is None and j > month_col_indices[-1]:
            f3_col = j
        elif cl in ("F1", "T1", "CURRENT MTD") and f1_col is None and j > month_col_indices[-1]:
            f1_col = j
        elif "confidence" in cl.lower() and conf_col is None:
            conf_col = j

    # Position-based fallback if header labels are missing
    if f12_col is None:
        f12_col = month_col_indices[-1] + 2
    if f6_col is None:
        f6_col = f12_col + 1
    if f3_col is None:
        f3_col = f6_col + 1
    if f1_col is None:
        f1_col = f3_col + 1

    # ── Step 3: Parse data rows ────────────────────────────────────────────
    line_items = []
    prev_category = None

    for row in all_rows[header_row_idx + 1:]:
        if not any(v is not None and v != "" for v in row):
            continue

        cat_val  = row[cat_col]  if cat_col  < len(row) else None
        item_val = row[item_col] if item_col < len(row) else None

        if item_val is None or item_val is False or isinstance(item_val, (int, float, datetime)):
            continue
        item_str = str(item_val).strip()
        if not item_str or len(item_str) < 2:
            continue

        if cat_val and isinstance(cat_val, str) and cat_val.strip():
            prev_category = cat_val.strip()
        cat_str = prev_category or ""

        monthly = [
            _safe_float(row[ci]) if ci < len(row) else None
            for ci in month_col_indices
        ]
        f12 = _safe_float(row[f12_col]) if f12_col < len(row) else None
        f6  = _safe_float(row[f6_col])  if f6_col  < len(row) else None
        f3  = _safe_float(row[f3_col])  if f3_col  < len(row) else None
        f1  = _safe_float(row[f1_col])  if f1_col  < len(row) else None
        conf = _safe_float(row[conf_col]) if (conf_col and conf_col < len(row)) else None

        is_sub = (cat_val is None or str(cat_val).strip() == "")

        line_items.append({
            "category":    cat_str,
            "line_item":   item_str,
            "monthly":     monthly,
            # Mirror T12 parser's key names so the financials view doesn't
            # need to special-case "F12" vs "T12":
            "t12":         f12,
            "t6":          f6,
            "t3":          f3,
            "t1":          f1,
            "confidence":  conf,
            "is_revenue":  _is_revenue(cat_str, item_str),
            "is_expense":  _is_expense(cat_str, item_str),
            "is_noi":      _is_noi(item_str),
            "is_subtotal": is_sub,
        })

    if not line_items:
        errors.append("No line items could be parsed from this budget file.")
        return {"errors": errors, "warnings": warnings}

    # ── Step 4: Build summary ──────────────────────────────────────────────
    def find_value(keyword: str, col: str = "t12"):
        """Find first line-item containing keyword and return its value. Treats
        zero/None on a subtotal row as 'missing' so callers fall through to
        derive-by-summing. Filled-in subtotals (non-zero) are returned."""
        kw = keyword.lower()
        for li in line_items:
            if kw in li["line_item"].lower():
                v = li.get(col)
                if v is not None and v != 0:
                    return v
        return None

    total_revenue_f12  = find_value("total revenue", "t12") or find_value("total income", "t12")
    total_expenses_f12 = find_value("total operating expenses", "t12") or find_value("operating expenses", "t12")
    noi_f12 = find_value("net operating income", "t12") or find_value("noi", "t12")

    # Derive by sum if subtotals not present or are zero
    def _sum_col(items, col):
        s = sum((li.get(col) or 0) for li in items)
        return s if s != 0 else None

    if total_revenue_f12 is None:
        total_revenue_f12 = _sum_col(
            [li for li in line_items if li["is_revenue"] and not li["is_subtotal"]],
            "t12",
        )
    if total_expenses_f12 is None:
        total_expenses_f12 = _sum_col(
            [li for li in line_items if li["is_expense"] and not li["is_subtotal"]],
            "t12",
        )
    if noi_f12 is None and total_revenue_f12 is not None and total_expenses_f12 is not None:
        noi_f12 = total_revenue_f12 - total_expenses_f12

    noi_margin = (
        noi_f12 / total_revenue_f12
        if (noi_f12 is not None and total_revenue_f12)
        else None
    )

    # T6/T3/T1 totals — same derivation pattern
    def _period_total_for(keyword_list, col):
        for kw in keyword_list:
            v = find_value(kw, col)
            if v is not None:
                return v
        return None

    def _derived(predicate, col):
        return _sum_col([li for li in line_items if predicate(li)], col)

    is_rev_item = lambda li: li["is_revenue"] and not li["is_subtotal"]
    is_exp_item = lambda li: li["is_expense"] and not li["is_subtotal"]

    summary = {
        "total_revenue_t12":   total_revenue_f12,
        "total_expenses_t12":  total_expenses_f12,
        "noi_t12":             noi_f12,
        "noi_margin_t12":      noi_margin,
    }
    for period in ("t6", "t3", "t1"):
        rev = _period_total_for(["total revenue", "total income"], period) or _derived(is_rev_item, period)
        exp = _period_total_for(["total operating expenses", "operating expenses"], period) or _derived(is_exp_item, period)
        noi = _period_total_for(["net operating income", "noi"], period)
        if noi is None and rev is not None and exp is not None:
            noi = rev - exp
        summary[f"total_revenue_{period}"]  = rev
        summary[f"total_expenses_{period}"] = exp
        summary[f"noi_{period}"]            = noi

    # ── Step 5: Monthly totals ─────────────────────────────────────────────
    def monthly_sum(predicate):
        out = [None] * len(month_dates)
        for li in line_items:
            if not predicate(li):
                continue
            for i, m in enumerate(li["monthly"]):
                if m is None:
                    continue
                out[i] = (out[i] or 0) + m
        return out

    monthly_revenue  = monthly_sum(lambda li: li["is_revenue"] and not li["is_subtotal"])
    monthly_expenses = monthly_sum(lambda li: li["is_expense"] and not li["is_subtotal"])
    monthly_noi = [
        ((r or 0) - (e or 0)) if (r is not None or e is not None) else None
        for r, e in zip(monthly_revenue, monthly_expenses)
    ]

    return {
        "as_of_date":     as_of_date,
        "month_dates":    month_dates,
        "line_items":     line_items,
        "summary":        summary,
        "monthly_totals": {
            "revenue":  monthly_revenue,
            "expenses": monthly_expenses,
            "noi":      monthly_noi,
        },
        # Convenience aliases used by views/financials.py
        "monthly_revenue": monthly_revenue,
        "annual_revenue":  summary["total_revenue_t12"],
        "errors":   errors,
        "warnings": warnings,
    }
